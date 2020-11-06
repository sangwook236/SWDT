#!/usr/bin/env python
# -*- coding: UTF-8 -*-

# REF [site] >> https://foss.heptapod.net/openpyxl/openpyxl

import openpyxl

# REF [site] >> https://foss.heptapod.net/openpyxl/openpyxl
def basic_example():
	wb = openpyxl.Workbook()

	# Grab the active worksheet.
	ws = wb.active

	# Data can be assigned directly to cells.
	ws['A1'] = 42

	# Rows can also be appended.
	ws.append([1, 2, 3])

	# Python types will automatically be converted.
	import datetime
	ws['A2'] = datetime.datetime.now()

	# Save the file.
	wb.save('./sample.xlsx')

def main():
	basic_example()

#--------------------------------------------------------------------

if '__main__' == __name__:
	main()
